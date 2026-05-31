from django.shortcuts import (
    render,
    redirect,
    get_object_or_404
)

from django.http import HttpResponse

from django.utils import timezone

from django.contrib.admin.views.decorators import (
    staff_member_required
)

from django.db.models import (
    Q,
    Count
)

import openpyxl

from reportlab.lib import colors

from reportlab.lib.pagesizes import (
    A4,
    landscape
)

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib.styles import (
    getSampleStyleSheet
)

from .models import (
    Notification
)

from recruitment.models import (
    Application,
    Position
)

from recruitment.forms import (
    PositionForm
)



@staff_member_required
def recruiter_dashboard(request):

    # RECENT APPLICATIONS

    applications = Application.objects.select_related(
        'candidate',
        'position'
    ).order_by('-created_at')[:5]

    # TOTAL COUNTS

    total_count = Application.objects.count()

    pending_count = Application.objects.filter(
        status='pending'
    ).count()

    shortlisted_count = Application.objects.filter(
        status='shortlisted'
    ).count()

    interview_count = Application.objects.filter(
        status='interview'
    ).count()

    accepted_count = Application.objects.filter(
        status='accepted'
    ).count()

    rejected_count = Application.objects.filter(
        status='rejected'
    ).count()

    # ACTIVE POSITIONS

    active_positions = Position.objects.filter(
        is_active=True
    ).count()

    # ACCEPTANCE RATE

    if total_count > 0:

        acceptance_rate = round(
            (accepted_count / total_count) * 100
        )

    else:

        acceptance_rate = 0

    # TOP UNIVERSITIES

    top_universities = Application.objects.values(
        'candidate__institution'
    ).annotate(
        total=Count('id')
    ).order_by('-total')[:5]

    # POPULAR POSITIONS

    popular_positions = Application.objects.values(
        'position__title'
    ).annotate(
        total=Count('id')
    ).order_by('-total')[:5]

    return render(
        request,
        'recruiter/dashboard.html',
        {

            'applications': applications,

            'total_count': total_count,
            'pending_count': pending_count,
            'shortlisted_count': shortlisted_count,
            'interview_count': interview_count,
            'accepted_count': accepted_count,
            'rejected_count': rejected_count,

            'active_positions': active_positions,
            'acceptance_rate': acceptance_rate,

            'top_universities': top_universities,
            'popular_positions': popular_positions,

            'notifications': Notification.objects.order_by(
                '-created_at'
            )[:10],

            'unread_notifications_count': Notification.objects.filter(
                is_read=False
            ).count(),

        }
    )



@staff_member_required
def recruiter_positions(request):

    positions = Position.objects.all().order_by(
        '-created_at'
    )

    return render(
        request,
        'recruiter/positions.html',
        {
            'positions': positions,

            'notifications': Notification.objects.order_by(
                '-created_at'
            )[:10],

            'unread_notifications_count': Notification.objects.filter(
                is_read=False
            ).count(),
        }
    )

@staff_member_required
def recruiter_create_position(request):

    form = PositionForm()

    if request.method == 'POST':

        form = PositionForm(request.POST)

        if form.is_valid():

            form.save()

            return redirect(
                '/recruiter/positions/'
            )

    return render(
        request,
        'recruiter/create_position.html',
        {
            'form': form
        }
    )

@staff_member_required
def recruiter_edit_position(request, id):

    position = get_object_or_404(
        Position,
        id=id
    )

    form = PositionForm(
        instance=position
    )

    if request.method == 'POST':

        form = PositionForm(
            request.POST,
            instance=position
        )

        if form.is_valid():

            form.save()

            return redirect(
                '/recruiter/positions/'
            )

    return render(
        request,
        'recruiter/edit_position.html',
        {
            'form': form,
            'position': position
        }
    )

@staff_member_required
def recruiter_delete_position(request, id):

    position = get_object_or_404(
        Position,
        id=id
    )

    position.delete()

    return redirect(
        '/recruiter/positions/'
    )



@staff_member_required
def recruiter_candidates(request):

    applications = Application.objects.select_related(
        'candidate',
        'position'
    ).order_by('-created_at')

    # SEARCH

    search = request.GET.get('search')

    if search:

        applications = applications.filter(

            Q(candidate__first_name__icontains=search) |

            Q(candidate__last_name__icontains=search) |

            Q(candidate__institution__icontains=search) |

            Q(candidate__scientific_field__icontains=search)

        )

    # STATUS FILTER

    status = request.GET.get('status')

    if status:

        applications = applications.filter(
            status=status
        )

    return render(
        request,
        'recruiter/candidates.html',
        {
            'applications': applications,
            'search': search,
            'status': status,

            'notifications': Notification.objects.order_by(
                '-created_at'
            )[:10],

            'unread_notifications_count': Notification.objects.filter(
                is_read=False
            ).count(),
        }
    )


@staff_member_required
def recruiter_applications(request):

    applications = Application.objects.select_related(
        'candidate',
        'position'
    ).order_by('-created_at')

    # SEARCH

    search = request.GET.get('search')

    if search:

        applications = applications.filter(

            Q(candidate__first_name__icontains=search) |

            Q(candidate__last_name__icontains=search) |

            Q(candidate__institution__icontains=search) |

            Q(position__title__icontains=search)

        )

    # FILTER STATUS

    status = request.GET.get('status')

    if status:

        applications = applications.filter(
            status=status
        )

    # COUNTS

    total_count = Application.objects.count()

    pending_count = Application.objects.filter(
        status='pending'
    ).count()

    shortlisted_count = Application.objects.filter(
        status='shortlisted'
    ).count()

    interview_count = Application.objects.filter(
        status='interview'
    ).count()

    accepted_count = Application.objects.filter(
        status='accepted'
    ).count()

    rejected_count = Application.objects.filter(
        status='rejected'
    ).count()

    return render(
        request,
        'recruiter/applications.html',
        {
            'applications': applications,

            'total_count': total_count,
            'pending_count': pending_count,
            'shortlisted_count': shortlisted_count,
            'interview_count': interview_count,
            'accepted_count': accepted_count,
            'rejected_count': rejected_count,

            'search': search,
            'status': status,

            'notifications': Notification.objects.order_by(
                '-created_at'
            )[:10],

            'unread_notifications_count': Notification.objects.filter(
                is_read=False
            ).count(),
        }
    )


def export_applications_excel(request):

    workbook = openpyxl.Workbook()

    worksheet = workbook.active
    worksheet.title = "Applicants"

    headers = [

        "Nama Lengkap",
        "Username",
        "Email",
        "WhatsApp",
        "Kota",
        "Gender",
        "Status Kandidat",
        "Institusi",
        "Program Studi",
        "Pekerjaan",
        "Bidang Ilmu",
        "Skill OJS",
        "Posisi Dilamar",
        "Status Lamaran",
        "Jurnal Ditugaskan",
        "Interviewer",
        "Tanggal Interview",
        "Jam Interview",
        "Tanggal Melamar",
        "CV",
        "Portfolio",

    ]

    # Header Style
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill(
        start_color="163B9F",
        end_color="163B9F",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for col_num, header in enumerate(headers, start=1):

        cell = worksheet.cell(
            row=1,
            column=col_num
        )

        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    applications = Application.objects.select_related(
        'candidate',
        'candidate__user',
        'position'
    )

    for row_num, application in enumerate(
        applications,
        start=2
    ):

        candidate = application.candidate

        # Nama Lengkap
        full_name = (
            f"{candidate.first_name} {candidate.last_name}"
        ).strip()

        worksheet.cell(
            row=row_num,
            column=1
        ).value = full_name

        # Username
        worksheet.cell(
            row=row_num,
            column=2
        ).value = candidate.user.username

        # Email
        worksheet.cell(
            row=row_num,
            column=3
        ).value = candidate.user.email

        # WhatsApp
        worksheet.cell(
            row=row_num,
            column=4
        ).value = candidate.phone

        # Kota
        worksheet.cell(
            row=row_num,
            column=5
        ).value = candidate.city

        # Gender
        worksheet.cell(
            row=row_num,
            column=6
        ).value = candidate.gender

        # Status Kandidat
        worksheet.cell(
            row=row_num,
            column=7
        ).value = candidate.get_status_display()

        # Institusi
        worksheet.cell(
            row=row_num,
            column=8
        ).value = candidate.institution

        # Program Studi
        worksheet.cell(
            row=row_num,
            column=9
        ).value = candidate.study_program

        # Pekerjaan
        worksheet.cell(
            row=row_num,
            column=10
        ).value = candidate.occupation

        # Bidang Ilmu
        worksheet.cell(
            row=row_num,
            column=11
        ).value = candidate.scientific_field

        # Skill OJS
        worksheet.cell(
            row=row_num,
            column=12
        ).value = candidate.get_ojs_skill_display()

        # Posisi Dilamar
        worksheet.cell(
            row=row_num,
            column=13
        ).value = application.position.title

        # Status Lamaran
        worksheet.cell(
            row=row_num,
            column=14
        ).value = application.get_status_display()

        # Jurnal Ditugaskan
        worksheet.cell(
            row=row_num,
            column=15
        ).value = application.assigned_journal

        # Interviewer
        worksheet.cell(
            row=row_num,
            column=16
        ).value = getattr(
            application,
            'assigned_interviewer',
            ''
        )

        # Tanggal Interview
        worksheet.cell(
            row=row_num,
            column=17
        ).value = application.interview_date

        # Jam Interview
        worksheet.cell(
            row=row_num,
            column=18
        ).value = (
            application.interview_time.strftime("%H:%M")
            if application.interview_time
            else ''
        )

        # Tanggal Melamar
        worksheet.cell(
            row=row_num,
            column=19
        ).value = application.created_at.strftime(
            "%d-%m-%Y %H:%M"
        )

        # CV
        worksheet.cell(
            row=row_num,
            column=20
        ).value = request.build_absolute_uri(
            candidate.cv.url
        )

        # Portfolio
        worksheet.cell(
            row=row_num,
            column=21
        ).value = (
            request.build_absolute_uri(
                candidate.portfolio.url
            )
            if candidate.portfolio
            else ''
        )

    # Auto Width Column
    for column in worksheet.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:

                if cell.value:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            except Exception:

                pass

        worksheet.column_dimensions[
            column_letter
        ].width = max_length + 5

    response = HttpResponse(
        content_type=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response[
        'Content-Disposition'
    ] = (
        'attachment; '
        'filename=litera_applicants.xlsx'
    )

    workbook.save(response)

    return response


def export_applications_pdf(request):

    response = HttpResponse(
        content_type='application/pdf'
    )

    response[
        'Content-Disposition'
    ] = (
        'attachment; '
        'filename=litera_applicants.pdf'
    )

    document = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    elements = []

    styles = getSampleStyleSheet()

    # =====================================
    # TITLE
    # =====================================

    title = Paragraph(
        "LITERA PUBLISHING RECRUITMENT REPORT",
        styles['Title']
    )

    subtitle = Paragraph(
        f"Generated on {timezone.now().strftime('%d-%m-%Y %H:%M')}",
        styles['Normal']
    )

    elements.append(title)
    elements.append(subtitle)
    elements.append(Spacer(1, 15))

    # =====================================
    # DATA
    # =====================================

    applications = Application.objects.select_related(
        'candidate',
        'candidate__user',
        'position'
    )

    total = applications.count()

    accepted = applications.filter(
        status='accepted'
    ).count()

    rejected = applications.filter(
        status='rejected'
    ).count()

    interview = applications.filter(
        status='interview'
    ).count()

    shortlisted = applications.filter(
        status='shortlisted'
    ).count()

    summary = Paragraph(
        f"""
        <b>Total Applicants:</b> {total}
        &nbsp;&nbsp;&nbsp;&nbsp;
        <b>Accepted:</b> {accepted}
        &nbsp;&nbsp;&nbsp;&nbsp;
        <b>Rejected:</b> {rejected}
        &nbsp;&nbsp;&nbsp;&nbsp;
        <b>Interview:</b> {interview}
        &nbsp;&nbsp;&nbsp;&nbsp;
        <b>Shortlisted:</b> {shortlisted}
        """,
        styles['Normal']
    )

    elements.append(summary)
    elements.append(Spacer(1, 15))

    # =====================================
    # TABLE HEADER
    # =====================================

    data = [[

        'No',
        'Name',
        'Gender',
        'Institution',
        'Phone',
        'Position',
        'Status',
        'Journal Assign'

    ]]

    # =====================================
    # TABLE CONTENT
    # =====================================
    normal_style = styles['BodyText']

    for no, app in enumerate(applications, start=1):

        candidate = app.candidate

        full_name = (
            f"{candidate.first_name} {candidate.last_name}"
        ).strip()

        data.append([

            str(no),

            Paragraph(
                full_name,
                normal_style
            ),

            candidate.gender or '-',

            Paragraph(
                candidate.institution or '-',
                normal_style
            ),

            candidate.phone or '-',

            Paragraph(
                app.position.title,
                normal_style
            ),

            app.get_status_display(),

            Paragraph(
                app.assigned_journal or '-',
                normal_style
            ),

        ])

    # =====================================
    # TABLE
    # =====================================

    table = Table(

        data,

        colWidths=[

            25,   # No
            120,  # Nama
            50,   # Gender
            180,  # Institusi
            80,   # HP
            120,  # Posisi
            70,   # Status
            180,  # Jurnal

        ]

    )

    table.setStyle(
        TableStyle([

            (
                'BACKGROUND',
                (0, 0),
                (-1, 0),
                colors.HexColor('#163B9F')
            ),

            (
                'TEXTCOLOR',
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                'FONTNAME',
                (0, 0),
                (-1, 0),
                'Helvetica-Bold'
            ),

            (
                'FONTSIZE',
                (0, 0),
                (-1, -1),
                8
            ),

            (
                'GRID',
                (0, 0),
                (-1, -1),
                0.5,
                colors.black
            ),

            (
                'ROWBACKGROUNDS',
                (0, 1),
                (-1, -1),
                [
                    colors.whitesmoke,
                    colors.lightgrey
                ]
            ),

            (
                'ALIGN',
                (0, 0),
                (-1, -1),
                'CENTER'
            ),

            (
                'VALIGN',
                (0, 0),
                (-1, -1),
                'MIDDLE'
            ),

        ])
    )

    elements.append(table)

    elements.append(
        Spacer(1, 15)
    )

    footer = Paragraph(
        "PT Litera Integra Nusantara - Recruitment Management System",
        styles['Italic']
    )

    elements.append(footer)

    document.build(elements)

    return response